"""
create_entra_users_from_lw.py
LINE WORKS メンバーExcelから Entra ID（Azure AD）アカウントを一括作成するスクリプト。

■ 処理対象
  LWメンバー  : SNS_ID が空欄 かつ フリガナあり（実在メンバーのみ）
                  → SNS_ID 入り = M365あり = Entra ID 既存 → スキップ
                  → フリガナ空欄 = 共有LWアカウント（事業所・予定等）→ スキップ
  共有端末     : shared_devices.json に定義（スマホ・タブレット）

■ 使い方
  # ドライラン（作成せず確認だけ）
  python create_entra_users_from_lw.py --xlsx members.xlsx --dry-run

  # 実際に作成
  python create_entra_users_from_lw.py --xlsx members.xlsx

■ 必要な環境変数（.env）
  ENTRA_TENANT_ID       テナントID
  ENTRA_CLIENT_ID       アプリ登録 Client ID（User.ReadWrite.All 権限）
  ENTRA_CLIENT_SECRET   クライアントシークレット
  ENTRA_DOMAIN          UPN ドメイン（例: tseg7421.onmicrosoft.com）
  ENTRA_INIT_PASSWORD   初期パスワード（大小英数記号8文字以上）
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv

load_dotenv()

# ── 設定 ─────────────────────────────────────────────────────────────────────
TENANT_ID  = os.getenv("ENTRA_TENANT_ID", "")
CLIENT_ID  = os.getenv("ENTRA_CLIENT_ID", "")
CLIENT_SEC = os.getenv("ENTRA_CLIENT_SECRET", "")
DOMAIN     = os.getenv("ENTRA_DOMAIN", "tseg7421.onmicrosoft.com")
INIT_PASS  = os.getenv("ENTRA_INIT_PASSWORD", "")

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
TOKEN_URL  = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"

DEVICES_JSON = Path(__file__).parent / "shared_devices.json"

# Excel 列名
COL_SEI     = "姓"
COL_MEI     = "名"
COL_ID      = "ID"
COL_KANA    = "姓(フリガナ)"   # 空欄 = 共有LWアカウント → 除外
COL_SNS_ID  = "SNS_ID"         # 入力済み = M365あり → 除外


# ── Graph API ────────────────────────────────────────────────────────────────
def _get_token() -> str:
    r = requests.post(TOKEN_URL, data={
        "grant_type":    "client_credentials",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SEC,
        "scope":         "https://graph.microsoft.com/.default",
    }, timeout=15)
    r.raise_for_status()
    return r.json()["access_token"]


def _get_existing_upns(token: str) -> set[str]:
    """既存 Entra ID ユーザーの UPN セットを返す。"""
    upns: set[str] = set()
    url = f"{GRAPH_BASE}/users?$select=userPrincipalName&$top=999"
    headers = {"Authorization": f"Bearer {token}"}
    while url:
        r = requests.get(url, headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json()
        for u in data.get("value", []):
            upns.add(u["userPrincipalName"].lower())
        url = data.get("@odata.nextLink")
    return upns


def _create_user(token: str, display_name: str, given: str, sur: str, upn: str) -> dict:
    nick = upn.split("@")[0]
    body: dict = {
        "accountEnabled": True,
        "displayName":    display_name,
        "userPrincipalName": upn,
        "mailNickname":   nick,
        "passwordProfile": {
            "password":                      INIT_PASS,
            "forceChangePasswordNextSignIn": True,
        },
    }
    # givenName / surname は空文字だと Graph API がエラーを返すため、値がある場合のみ追加
    if given:
        body["givenName"] = given
    if sur:
        body["surname"] = sur
    r = requests.post(
        f"{GRAPH_BASE}/users",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json=body,
        timeout=15,
    )
    return {"ok": r.ok, "status": r.status_code,
            "detail": r.json() if not r.ok else {}}


# ── LW ID → UPN 変換 ─────────────────────────────────────────────────────────
def _lw_id_to_upn(lw_id: str) -> str:
    # wo.37995 → wo37995  /  それ以外はそのまま
    normalized = re.sub(r'^wo\.(\d+)$', r'wo\1', lw_id, flags=re.IGNORECASE)
    return f"{normalized}@{DOMAIN}"


# ── Excel 読み込み ────────────────────────────────────────────────────────────
def _load_members(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    def idx(col: str) -> int:
        return headers.index(col)

    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sei    = (row[idx(COL_SEI)]    or "").strip()
        mei    = (row[idx(COL_MEI)]    or "").strip()
        lw_id  = (row[idx(COL_ID)]     or "").strip()
        kana   = (row[idx(COL_KANA)]   or "").strip()
        sns_id = (row[idx(COL_SNS_ID)] or "").strip()
        if not sei:
            continue
        members.append({
            "sei": sei, "mei": mei, "lw_id": lw_id,
            "kana": kana, "sns_id": sns_id,
        })
    return members


# ── メイン ────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx",    required=True, help="LW メンバー Excel パス")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.dry_run:
        missing = [k for k, v in {
            "ENTRA_TENANT_ID":     TENANT_ID,
            "ENTRA_CLIENT_ID":     CLIENT_ID,
            "ENTRA_CLIENT_SECRET": CLIENT_SEC,
            "ENTRA_INIT_PASSWORD": INIT_PASS,
        }.items() if not v]
        if missing:
            print(f"[ERROR] .env 未設定: {', '.join(missing)}")
            sys.exit(1)
    # ドライランは .env チェックをスキップ

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"[ERROR] ファイルが見つかりません: {xlsx_path}")
        sys.exit(1)

    all_members = _load_members(xlsx_path)

    # ── 分類 ──────────────────────────────────────────────────────────────────
    skip_shared = [m for m in all_members if not m["kana"]]          # 共有LWアカウント
    skip_m365   = [m for m in all_members if m["kana"] and m["sns_id"]]  # M365あり
    targets_lw  = [m for m in all_members if m["kana"] and not m["sns_id"]]  # 作成対象

    # 共有端末
    devices = []
    if DEVICES_JSON.exists():
        devices = json.loads(DEVICES_JSON.read_text(encoding="utf-8"))

    print(f"LW メンバー総数  : {len(all_members)} 名")
    print(f"  共有LWアカウント除外: {len(skip_shared)} 件"
          f"  ({', '.join(m['sei']+m['mei'] for m in skip_shared)})")
    print(f"  M365あり（スキップ）: {len(skip_m365)} 名")
    print(f"  Entra ID 作成対象  : {len(targets_lw)} 名")
    print(f"共有端末           : {len(devices)} 台")
    print()

    # ── 作成予定一覧 ──────────────────────────────────────────────────────────
    print("─" * 65)
    print(f"{'区分':<8} {'氏名':<12} {'LW ID / 端末名':<32} {'作成予定 UPN'}")
    print("─" * 65)
    for m in targets_lw:
        upn = _lw_id_to_upn(m["lw_id"])
        print(f"{'メンバー':<8} {m['sei']+m['mei']:<12} {m['lw_id']:<32} {upn}")
    for d in devices:
        print(f"{'端末':<8} {d['display_name']:<12} {'':32} {d['upn']}")
    print("─" * 65)

    if args.dry_run:
        print("\n[DRY RUN] 上記を作成します。--dry-run を外すと実際に作成されます。")
        return

    # ── Entra ID 処理 ─────────────────────────────────────────────────────────
    print("\nAzure 認証中...")
    token = _get_token()
    print("既存 Entra ID ユーザーを取得中...")
    existing = _get_existing_upns(token)
    print(f"既存: {len(existing)} 件\n")

    created, skipped, errors = [], [], []

    def _process(display: str, given: str, sur: str, upn: str, label: str) -> None:
        if upn.lower() in existing:
            print(f"  [SKIP] 既存: {upn}")
            skipped.append(upn)
            return
        result = _create_user(token, display, given, sur, upn)
        if result["ok"]:
            print(f"  [OK]   {label}: {upn}")
            created.append(upn)
        else:
            msg = result["detail"].get("error", {}).get("message", str(result["detail"]))
            print(f"  [ERR]  {label}: {upn} → {msg}")
            errors.append((upn, msg))
        time.sleep(0.3)

    print("── LW メンバー ──")
    for m in targets_lw:
        upn = _lw_id_to_upn(m["lw_id"])
        _process(f"{m['sei']} {m['mei']}", m["mei"], m["sei"], upn, m["sei"]+m["mei"])

    print("\n── 共有端末 ──")
    for d in devices:
        _process(d["display_name"], "", "", d["upn"], d["display_name"])

    print(f"\n{'='*60}")
    print(f"作成完了    : {len(created)} 件")
    print(f"既存スキップ: {len(skipped)} 件")
    print(f"エラー      : {len(errors)} 件")
    if errors:
        for upn, msg in errors:
            print(f"  {upn}: {msg}")
    print(f"\n初期パスワード: {INIT_PASS}")
    print("初回ログイン時にパスワード変更が必要です。")


if __name__ == "__main__":
    main()
