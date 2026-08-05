"""高千穂見積工数建値表をBlobへスナップショット(AI Q&Aの建値回答用)。

最新の価格表シート(シート名の日付6桁が最大のもの)を行単位のテキストに変換し、
lw-raw/tatene.json に保存する。AI Q&A(見積メンバー限定)が「○○の建値は？」
「標準工数は？」に、この表を根拠に答える。

実行(デスクトップ):
    python tools/export_tatene.py            # 取り込み
    python tools/export_tatene.py --dry-run  # 変換結果の確認のみ
    python tools/export_tatene.py --sheet "改定高千穂価格表251101"  # シート指定
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass

JST = timezone(timedelta(hours=9))
TATENE_PATH = Path(os.environ.get(
    "TATENE_PATH",
    r"Z:\takachiho\1_原紙・マスタデータフォルダ\12_事業本部共通\121_顧客関連"
    r"\1217_高千穂見積工数建値表.xlsx"))
BLOB_NAME = "tatene.json"


def _pick_sheet(wb) -> str:
    """シート名の6桁日付が最大のシート=最新版を選ぶ。無ければ最後のシート。"""
    best, best_key = None, -1
    for name in wb.sheetnames:
        m = re.search(r"(\d{6})", name)
        if m and int(m.group(1)) > best_key:
            best, best_key = name, int(m.group(1))
    return best or wb.sheetnames[-1]


def main() -> int:
    ap = argparse.ArgumentParser(description="建値表のスナップショット")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--sheet", default="", help="シート名を明示指定(既定は最新日付)")
    args = ap.parse_args()

    if not TATENE_PATH.exists():
        print(f"[ERROR] 建値表が見つかりません: {TATENE_PATH}", file=sys.stderr)
        return 1

    import openpyxl
    wb = openpyxl.load_workbook(TATENE_PATH, read_only=True, data_only=True)
    sheet = args.sheet or _pick_sheet(wb)
    if sheet not in wb.sheetnames:
        print(f"[ERROR] シートがありません: {sheet}", file=sys.stderr)
        return 1
    ws = wb[sheet]

    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        cells = []
        for v in row:
            if v is None:
                continue
            if isinstance(v, float) and v == int(v):
                v = int(v)
            s = str(v).strip()
            if s:
                cells.append(s)
        line = " / ".join(cells)
        if len(line) >= 3:
            lines.append(line[:200])
    wb.close()

    print(f"[SCAN] シート「{sheet}」→ {len(lines)} 行")
    if args.dry_run:
        for ln in lines[:15]:
            print(f"  {ln}")
        print("[DRY-RUN] Blobへは書き込みません")
        return 0

    payload = {
        "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        "source": str(TATENE_PATH),
        "sheet": sheet,
        "count": len(lines),
        "lines": lines,
    }
    conn = os.environ.get("AZURE_BLOB_CONNECTION_STRING", "")
    if not conn:
        print("ERROR: AZURE_BLOB_CONNECTION_STRING が未設定です", file=sys.stderr)
        return 1
    from azure.storage.blob import BlobServiceClient
    BlobServiceClient.from_connection_string(conn) \
        .get_blob_client(os.environ.get("LW_BLOB_CONTAINER", "lw-raw"), BLOB_NAME) \
        .upload_blob(json.dumps(payload, ensure_ascii=False, indent=1).encode("utf-8"),
                     overwrite=True)
    print(f"[DONE] 建値表を更新: lw-raw/{BLOB_NAME} (シート: {sheet})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
