"""動治工具・測定具・消耗品リスト(寄居・綾瀬).xlsx から一覧を抽出し、
tools_inventory.json を lw-raw Blob に出力する。

目的:
    本番の検索アプリ(Azure App Service)は社内NAS(Z:)を直接見られないため、
    工具リストを JSON 化して Blob 経由で渡す。アプリはこれを読んで
    「動治工具・測定具・消耗品検索」を提供する。
    （部品在庫 export_parts_inventory.py と同じ方式）

実行:
    Z: が見える環境(デスクトップ)で、リストが更新されたら実行する。
    python export_tools_inventory.py

入力 Excel(既定; 環境変数 TOOLS_XLSX_DIR で親フォルダを上書き可):
    Z:\takachiho\2to9_業務別フォルダ\94_動治工具・測定具・消耗品\
        941_動治工具・測定具・消耗品リスト\
            9411_動治工具・測定具・消耗品リスト(寄居).xlsx
            9412_動治工具・測定具・消耗品リスト(綾瀬).xlsx

出力 Blob:
    lw-raw/tools_inventory.json
    形式: {"generated_at": ISO8601, "count": N,
           "sites": ["寄居","綾瀬"], "categories": [...],
           "items": [{"site","cat","name","model","spec","maker","qty","unit"}, ...]}
    ※仕入先・単価は社外に出したくないノウハウのため出力しない。

シート構成の特徴:
    1行目=シート見出し、2行目=ヘッダー、3行目以降=データ。
    シートごとに列の並びが違うため、位置ではなく「列名」で拾う。
    品名はグループの先頭行にだけ入っているので直前の値を引き継ぐ。

環境変数(.env):
    AZURE_BLOB_CONNECTION_STRING  Blob 接続文字列
    LW_BLOB_CONTAINER             コンテナ名(省略時: lw-raw)
    TOOLS_XLSX_DIR                xlsxの親フォルダ(省略時: 上記の既定)
"""
from __future__ import annotations

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).with_name(".env"), encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    print("[ERROR] pip install openpyxl が必要です", file=sys.stderr)
    sys.exit(1)

try:
    from azure.storage.blob import BlobServiceClient
except ImportError:
    print("[ERROR] pip install azure-storage-blob python-dotenv", file=sys.stderr)
    sys.exit(1)

BLOB_CONN_STR = os.environ.get("AZURE_BLOB_CONNECTION_STRING", "")
BLOB_CONTAINER = os.environ.get("LW_BLOB_CONTAINER", "lw-raw")
TOOLS_BLOB = "tools_inventory.json"

DEFAULT_DIR = (
    r"Z:\takachiho\2to9_業務別フォルダ\94_動治工具・測定具・消耗品"
    r"\941_動治工具・測定具・消耗品リスト"
)
XLSX_DIR = Path(os.environ.get("TOOLS_XLSX_DIR", DEFAULT_DIR))

FILES = [
    ("寄居", "9411_動治工具・測定具・消耗品リスト(寄居).xlsx"),
    ("綾瀬", "9412_動治工具・測定具・消耗品リスト(綾瀬).xlsx"),
]

# 一覧データではないシート（目次・表紙・除却済み・仕様一覧）は除外する
SKIP_SHEETS = {"目次", "表紙", "除却"}
SKIP_KEYWORDS = ("仕様一覧",)

# 列名の候補（シートによって表記ゆれがあるため複数指定）
COL_KEYS = {
    "name":  ("品名", "品　名", "品　名"),
    "model": ("型式", "型　式", "型　式"),
    "spec":  ("仕様・説明", "用途・仕様", "用途", "仕様", "説明"),
    "maker": ("メーカー", "ﾒｰｶｰ"),
    "qty":   ("個数", "数量"),
    "unit":  ("発注単位", "単位"),
}


def _s(v) -> str:
    """セル値を表示用文字列に。整数は小数点を落とす。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).replace("\n", " ").strip()


def _norm_header(v) -> str:
    """ヘッダー名から空白類を除いて比較しやすくする。"""
    return _s(v).replace(" ", "").replace("　", "")


def _find_header_row(ws, max_scan: int = 8):
    """『品名』を含む行をヘッダー行とみなし、(行番号, {項目: 列index}) を返す。"""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        cells = [_norm_header(c) for c in row]
        if not any(c in ("品名",) or c.startswith("品名") for c in cells):
            continue
        colmap: dict = {}
        for key, cands in COL_KEYS.items():
            wanted = [_norm_header(x) for x in cands]
            for idx, c in enumerate(cells):
                if not c:
                    continue
                if c in wanted or any(c.startswith(w) for w in wanted):
                    colmap.setdefault(key, idx)
                    break
        if "name" in colmap or "model" in colmap:
            return i, colmap
    return None, {}


def parse_tools(site: str, xlsx_path: Path) -> list[dict]:
    """1ファイル分の工具リストを読み、アイテムのリストを返す。"""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    items: list[dict] = []

    for sn in wb.sheetnames:
        name_norm = _norm_header(sn)
        if name_norm in SKIP_SHEETS or any(k in sn for k in SKIP_KEYWORDS):
            continue
        ws = wb[sn]
        header_row, colmap = _find_header_row(ws)
        if not header_row:
            continue

        last_name = ""
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            def g(key: str) -> str:
                idx = colmap.get(key)
                return _s(row[idx]) if idx is not None and idx < len(row) else ""

            nm = g("name")
            if nm:
                last_name = nm          # 品名はグループ先頭行だけ → 以降へ引き継ぐ
            model, spec = g("model"), g("spec")

            # 型式も仕様も空 → 空行・区切り行
            if not model and not spec:
                continue
            # 小計・合計行を除外
            if any(x in (nm + model) for x in ("小計", "合計", "総計")):
                continue

            items.append({
                "site":  site,                  # 拠点（寄居/綾瀬）
                "cat":   sn,                    # カテゴリ（シート名）
                "name":  nm or last_name,       # 品名
                "model": model,                 # 型式
                "spec":  spec,                  # 仕様・用途
                "maker": g("maker"),            # メーカー
                "qty":   g("qty"),              # 個数
                "unit":  g("unit"),             # 発注単位
            })
    return items


def main() -> None:
    if not BLOB_CONN_STR:
        print("[ERROR] AZURE_BLOB_CONNECTION_STRING が未設定です", file=sys.stderr)
        sys.exit(1)

    if not XLSX_DIR.exists():
        print(f"[ERROR] フォルダが見つかりません: {XLSX_DIR}", file=sys.stderr)
        print("        Z: 未接続の可能性。中止します(既存のBlobを保護)。", file=sys.stderr)
        sys.exit(2)

    items: list[dict] = []
    for site, fname in FILES:
        path = XLSX_DIR / fname
        if not path.exists():
            print(f"[WARN] 見つかりません（スキップ）: {fname}", file=sys.stderr)
            continue
        got = parse_tools(site, path)
        print(f"[OK] {site}: {len(got)} 件  ({fname})")
        items.extend(got)

    if not items:
        print("[ERROR] アイテムが0件。シート構成を確認してください。中止します。", file=sys.stderr)
        sys.exit(3)

    categories = sorted({it["cat"] for it in items})
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": len(items),
        "sites": sorted({it["site"] for it in items}),
        "categories": categories,
        "items": items,
    }

    svc = BlobServiceClient.from_connection_string(BLOB_CONN_STR)
    container = svc.get_container_client(BLOB_CONTAINER)
    container.upload_blob(
        TOOLS_BLOB,
        json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        overwrite=True,
        content_type="application/json",
    )
    print(
        f"[DONE] 工具リストをアップロード: {BLOB_CONTAINER}/{TOOLS_BLOB} "
        f"({len(items)} 件 / {len(categories)} カテゴリ)"
    )


if __name__ == "__main__":
    main()
