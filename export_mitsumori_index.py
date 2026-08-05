"""過去見積書(231_見積書)をAI検索インデックスに取り込む。

AI Q&Aが「○○社のNC1-110整備、過去いくらで出してる？」等の質問に、
過去見積の件名・日付・金額・ファイル所在を根拠に答えられるようにする。
回答できるのは見積作成メンバー(ai_qa.py側で質問者を制限)のみ。

方式: xlsx見積の全シート(枚数無制限)から明細本文を、先頭シートから宛先・件名・
      日付・合計金額をヒューリスティックに抽出。複数シートのブックは1シート=
      1ドキュメントに分割して photo-index に media_type="mitsumori" でupsert。
      旧xls・PDFは第2期(未対応。件数のみログ)。

実行(デスクトップ):
    python export_mitsumori_index.py --dry-run --limit 20  # 抽出結果の確認
    python export_mitsumori_index.py                       # 差分取り込み
    python export_mitsumori_index.py --full                # 全量作り直し
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).with_name(".env"), encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).resolve().parent))

ROOT = Path(os.environ.get(
    "MITSUMORI_DIR", r"Z:\takachiho\2to9_業務別フォルダ\23_顧客・納入先別\231_見積書"))
MEDIA_TYPE = "mitsumori"
STATE_PATH = Path(__file__).with_name("mitsumori_index_state.json")
BATCH = 100

RX_GYO = re.compile(r"^[あかさたなはまやらわ]行")  # 「あ行」「あ行(あいうえお)」の両対応
RX_WORKNO = re.compile(r"([A-Z]{0,4}\d{3,6}-\d{2})", re.IGNORECASE)
# シート名の6桁日付(例: ISS260220 / ISS260804 (2))。あれば見積日として採用
RX_SHEET_DATE = re.compile(r"^\D{0,10}(\d{6})(?:\s*\(\d+\))?$")
SHEET_BODY_MAX = 20000  # 1シート(=1ドキュメント)の本文上限

# ── 見積用語カタログ(明細から自動抽出→Blob mitsumori_terms.json) ─────────────
# 規程・現場カタログに続く4つ目の辞書。AI Q&Aのキーワード変換に注入される。
TERMS_COUNTS_PATH = Path(__file__).with_name("mitsumori_terms_counts.json")
TERMS_BLOB = "mitsumori_terms.json"
TERMS_MIN_COUNT = 3   # 明細は繰り返しが多いためノイズ除けに3回以上
RX_KATAKANA = re.compile(r"[ァ-ヴー]{3,}")
RX_MODEL = re.compile(r"[A-Za-z][A-Za-z\-]{0,8}[0-9][A-Za-z0-9\-]{0,14}")
RX_WORK = re.compile(
    r"[一-龠ァ-ヴーA-Za-z0-9]{1,8}"
    r"(?:交換|整備|修理|組立|組付|分解|洗浄|塗装|溶接|研磨|研削|調整|点検|測定|"
    r"芯出し|取付|取外し|加工|補修|清掃|据付|搬入|搬出|改造|更新)")
TERMS_STOP = {"ミツモリ", "ゴウケイ", "ショウヒゼイ"}


def _customer_from_path(rel: Path) -> str:
    parts = rel.parts
    if parts and RX_GYO.match(parts[0]):
        return parts[1] if len(parts) > 2 else ""
    return parts[0] if len(parts) > 1 else ""


def _scan_sheet(ws, want_header: bool) -> dict:
    """1シートを走査(400行×20列)して本文行とヘッダー情報(宛先・件名・日付・最大数値)を返す。"""
    atesaki = kenmei = date_s = ""
    max_num = 0.0
    label_next = False
    lines: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=400, max_col=20):
        row_texts: list[str] = []
        for c in row:
            v = c.value
            if v is None:
                continue
            if want_header and label_next and isinstance(v, str) and v.strip():
                cand = v.strip()
                # 表ヘッダー行(品名 数量 単価…)の隣接セルを件名と誤認しない
                if not re.fullmatch(r"数\s*量|単\s*価|金\s*額|単\s*位|備\s*考", cand):
                    kenmei = kenmei or cand[:60]
                label_next = False
            s = str(v).strip()
            if not s:
                continue
            if isinstance(v, str):
                row_texts.append(s)
            if want_header:
                if ((s.endswith("様") or s.endswith("御中"))
                        and not atesaki and 3 <= len(s) <= 40):
                    atesaki = s
                if isinstance(v, str) and re.fullmatch(r"件\s*名|工事名|品\s*名", s):
                    label_next = True  # 次の非空セルを件名とみなす
                if hasattr(v, "year") and not date_s:
                    try:
                        date_s = v.strftime("%Y-%m-%d")
                    except Exception:
                        pass
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    if v > max_num:
                        max_num = float(v)
        joined = " ".join(row_texts)
        if len(joined) >= 4:
            lines.append(joined)
    return {"lines": lines, "atesaki": atesaki, "kenmei": kenmei,
            "date": date_s, "max_num": max_num}


def _extract(p: Path) -> dict:
    """通常ブック用: 全シートから明細本文、先頭シートからヘッダー情報を抽出。

    シート数無制限・シートごと400行×20列・2000字・全体30万字まで。
    """
    import openpyxl
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        return _extract_normal(wb)
    finally:
        wb.close()


def _extract_normal(wb) -> dict:
    body_parts: list[str] = []
    hdr: dict = {}
    multi = len(wb.worksheets) > 1
    for si, ws in enumerate(wb.worksheets):
        info = _scan_sheet(ws, want_header=(si == 0))
        if si == 0:
            hdr = info
        if info["lines"]:
            sheet_text = "\n".join(info["lines"])[:2000]
            if multi:
                sheet_text = f"◆シート「{ws.title}」\n" + sheet_text
            body_parts.append(sheet_text)
        if sum(len(x) for x in body_parts) >= 300000:
            break
    body = "\n".join(body_parts)[:300000]
    return {"atesaki": hdr.get("atesaki", ""), "kenmei": hdr.get("kenmei", ""),
            "date": hdr.get("date", ""),
            "gokei": int(hdr.get("max_num", 0)) if hdr.get("max_num", 0) >= 1000 else 0,
            "body": body}


def _docs_for(p: Path, rel: Path, indexed_at: str) -> tuple[list[dict], str]:
    """1ファイル分の(検索ドキュメント群, 明細テキスト)を返す。

    複数シートのブックは1シート=1ドキュメントに分割する(台帳型・機械加工型とも)。
    見積日はシート名の6桁日付(ISS260220等) → シート内の日付 → ファイル更新日の
    順で採用。金額はそのシート内の最大数値。単一シートのブックは従来どおり
    1ファイル=1ドキュメント(IDも従来と同じ)。
    """
    import openpyxl
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    sheets: list[tuple[str, dict]] = []
    try:
        customer = _customer_from_path(rel)
        for ws in wb.worksheets:
            info = _scan_sheet(ws, want_header=True)
            if info["lines"]:
                sheets.append((ws.title, info))
    finally:
        wb.close()
    try:
        mtime_date = datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d")
    except OSError:
        mtime_date = ""
    multi = len(sheets) > 1
    docs: list[dict] = []
    bodies: list[str] = []
    for title, info in sheets:
        body = "\n".join(info["lines"])[:SHEET_BODY_MAX]
        m_d = RX_SHEET_DATE.match(title.strip())
        date_s = ""
        if m_d:
            try:
                date_s = datetime.strptime(
                    "20" + m_d.group(1), "%Y%m%d").strftime("%Y-%m-%d")
            except ValueError:
                date_s = ""
        date_s = date_s or info["date"] or mtime_date
        kenmei = info["kenmei"] or (f"{p.stem}◆{title}" if multi else p.stem)
        mw = (RX_WORKNO.search(p.stem) or RX_WORKNO.search(title)
              or RX_WORKNO.search(kenmei))
        workno = mw.group(1) if mw else ""
        gokei = int(info["max_num"]) if info["max_num"] >= 1000 else 0
        gs = f"{gokei:,}円" if gokei else "(金額抽出不可)"
        text = (f"【過去見積】顧客: {customer} / 宛先: {info['atesaki'] or '不明'} / "
                f"件名: {kenmei} / 見積日: {date_s or '不明'} / 合計金額: {gs}"
                + (f" / 工番: {workno}" if workno else "")
                + f"\nファイル: {p}" + (f" (シート: {title})" if multi else "")
                + (f"\n明細:\n{body}" if body else ""))
        id_src = f"mitsumori::{rel}::{title}" if multi else f"mitsumori::{rel}"
        docs.append({
            "id": hashlib.sha256(id_src.encode("utf-8")).hexdigest(),
            "file_path": str(p),
            "file_name": f"{p.name}◆{title}" if multi else p.name,
            "workno": workno,
            "workno_name": f"{customer}/{kenmei}"[:100],
            "phase": "",
            "media_type": MEDIA_TYPE,
            "capture_date": None,
            "capture_date_raw": date_s,
            "extension": p.suffix.lower(),
            "folder_path": str(p.parent),
            "indexed_at": indexed_at,
            "content_text": text,
        })
        bodies.append(body)
    return docs, "\n".join(bodies)[:300000]


def main() -> int:
    ap = argparse.ArgumentParser(description="過去見積のインデックス取り込み")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--limit", type=int, default=0, help="dry-run時の件数制限")
    ap.add_argument("--full", action="store_true", help="状態を無視して全量再取込")
    args = ap.parse_args()

    if not ROOT.is_dir():
        print(f"[ERROR] {ROOT}", file=sys.stderr)
        return 1

    state = {}
    if STATE_PATH.exists() and not args.full:
        state = json.loads(STATE_PATH.read_text(encoding="utf-8"))

    targets = []
    skipped_old = pdf_n = xls_n = 0
    current_keys = set()
    for dirpath, _, filenames in os.walk(ROOT):
        for fn in filenames:
            if fn.startswith("~$"):
                continue
            low = fn.lower()
            p = Path(dirpath) / fn
            if low.endswith(".pdf"):
                pdf_n += 1
                continue
            if low.endswith(".xls"):
                xls_n += 1
                continue
            if not low.endswith(".xlsx"):
                continue
            rel = p.relative_to(ROOT)
            key = str(rel)
            current_keys.add(key)
            try:
                mt = p.stat().st_mtime
            except OSError:
                continue
            prev = state.get(key)
            if (prev.get("mt") if isinstance(prev, dict) else prev) == mt:
                skipped_old += 1
                continue
            targets.append((p, rel, mt))

    removed = [k for k in state if k not in current_keys]
    print(f"[SCAN] 対象xlsx 新規/更新 {len(targets):,} 件 / 変更なし {skipped_old:,} 件 / "
          f"削除 {len(removed):,} 件 (未対応: 旧xls {xls_n:,}・PDF {pdf_n:,})")

    if args.dry_run:
        n = args.limit or 20
        indexed_at = datetime.now(timezone.utc).isoformat()
        for p, rel, _ in targets[:n]:
            try:
                ds, body = _docs_for(p, rel, indexed_at)
                print(f"--- {rel} → {len(ds)}ドキュメント")
                for d in ds[:3]:
                    print(f"    {d['content_text'].splitlines()[0][:130]}")
                if len(ds) > 3:
                    print(f"    …ほか{len(ds) - 3}シート")
            except Exception as e:
                print(f"--- {rel}\n    [失敗] {type(e).__name__}: {e}")
        return 0

    from azure.core.credentials import AzureKeyCredential
    from azure.search.documents import SearchClient
    from rag.config import SEARCH_INDEX_NAME, ensure_search_credentials

    endpoint, api_key = ensure_search_credentials()
    client = SearchClient(endpoint, SEARCH_INDEX_NAME, AzureKeyCredential(api_key))

    if args.full:
        old_ids = [r["id"] for r in client.search(
            search_text="*", filter=f"media_type eq '{MEDIA_TYPE}'",
            select=["id"], top=100000)]
        if old_ids:
            for i in range(0, len(old_ids), BATCH):
                client.delete_documents([{"id": x} for x in old_ids[i:i + BATCH]])
            print(f"[DELETE] 旧見積チャンク {len(old_ids):,} 件を削除")
        state = {}

    indexed_at = datetime.now(timezone.utc).isoformat()
    docs, ok, err, doc_total = [], 0, 0, 0
    processed_bodies: list[str] = []
    stale_ids: list[str] = []
    for p, rel, mt in targets:
        try:
            ds, body = _docs_for(p, rel, indexed_at)
            docs.extend(ds)
            doc_total += len(ds)
            if body:
                processed_bodies.append(body)
            # 前回登録したが今回消えたドキュメント(削除されたシート等)を削除対象に
            new_ids = {d["id"] for d in ds}
            prev = state.get(str(rel))
            prev_ids = (prev.get("ids", []) if isinstance(prev, dict)
                        else ([hashlib.sha256(f"mitsumori::{rel}".encode("utf-8"))
                               .hexdigest()] if prev else []))
            stale_ids += [i for i in prev_ids if i not in new_ids]
            state[str(rel)] = {"mt": mt, "ids": sorted(new_ids)}
            ok += 1
        except Exception as e:
            err += 1
            if err <= 10:
                print(f"[WARN] 抽出失敗 {rel}: {type(e).__name__}: {e}")
        if len(docs) >= BATCH:
            client.merge_or_upload_documents(docs)
            docs = []
            if ok % 500 < BATCH:
                print(f"  ... {ok:,} ファイル / {doc_total:,} ドキュメント登録")
    if docs:
        client.merge_or_upload_documents(docs)

    # 消されたシートの旧ドキュメントを削除
    if stale_ids:
        for i in range(0, len(stale_ids), BATCH):
            client.delete_documents([{"id": x} for x in stale_ids[i:i + BATCH]])
        print(f"[DELETE] 消えたシート等の旧ドキュメント {len(stale_ids):,} 件を削除")

    # 削除されたファイルをインデックスからも削除
    if removed:
        del_ids = []
        for k in removed:
            prev = state.get(k)
            if isinstance(prev, dict):
                del_ids += [{"id": i} for i in prev.get("ids", [])]
            else:
                del_ids.append({"id": hashlib.sha256(
                    f"mitsumori::{k}".encode()).hexdigest()})
        for i in range(0, len(del_ids), BATCH):
            client.delete_documents(del_ids[i:i + BATCH])
        for k in removed:
            state.pop(k, None)
        print(f"[DELETE] 削除ファイル {len(removed):,} 件分 "
              f"{len(del_ids):,} ドキュメントをインデックスから削除")

    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")

    # ── 見積用語カタログの更新(処理したファイルの明細から語彙を累積) ────────
    try:
        _update_terms_catalog(processed_bodies, full=args.full)
    except Exception as e:
        print(f"[WARN] 用語カタログ更新失敗: {e}")

    print(f"[DONE] 登録 {ok:,} ファイル ({doc_total:,} ドキュメント) / 失敗 {err:,} 件。"
          f"AI Q&A(見積メンバー限定)で過去見積に答えられます。")
    return 0


def _update_terms_catalog(bodies: list[str], full: bool = False) -> None:
    """明細テキストから語彙を抽出し、累積カウントを更新してBlobへ出力する。

    カウントはローカル(mitsumori_terms_counts.json)に累積。--full時は作り直し。
    出現TERMS_MIN_COUNT回以上の語だけをBlobのカタログに載せる(金額は含まない)。
    """
    from collections import Counter
    counts = {"カタカナ": {}, "型式": {}, "作業": {}}
    if TERMS_COUNTS_PATH.exists() and not full:
        counts = json.loads(TERMS_COUNTS_PATH.read_text(encoding="utf-8"))
    kata, model, work = (Counter(counts.get("カタカナ", {})),
                         Counter(counts.get("型式", {})),
                         Counter(counts.get("作業", {})))
    for body in bodies:
        for m in RX_KATAKANA.findall(body):
            if m not in TERMS_STOP:
                kata[m] += 1
        for m in RX_MODEL.findall(body):
            m = m.upper().rstrip("-")
            if len(m) >= 3:
                model[m] += 1
        for m in RX_WORK.findall(body):
            if not m.startswith(("の", "を", "は", "と", "が")):
                work[m] += 1
    TERMS_COUNTS_PATH.write_text(json.dumps(
        {"カタカナ": dict(kata), "型式": dict(model), "作業": dict(work)},
        ensure_ascii=False), encoding="utf-8")

    catalog = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "カタカナ": [w for w, c in kata.most_common(300) if c >= TERMS_MIN_COUNT],
        "型式": [w for w, c in model.most_common(300) if c >= TERMS_MIN_COUNT],
        "作業": [w for w, c in work.most_common(300) if c >= TERMS_MIN_COUNT],
    }
    conn = os.environ.get("AZURE_BLOB_CONNECTION_STRING", "")
    if not conn:
        print("[WARN] AZURE_BLOB_CONNECTION_STRING未設定のため用語カタログ未出力")
        return
    from azure.storage.blob import BlobServiceClient
    BlobServiceClient.from_connection_string(conn) \
        .get_blob_client(os.environ.get("LW_BLOB_CONTAINER", "lw-raw"), TERMS_BLOB) \
        .upload_blob(json.dumps(catalog, ensure_ascii=False, indent=1).encode("utf-8"),
                     overwrite=True)
    print(f"[TERMS] 見積用語カタログ更新: カタカナ{len(catalog['カタカナ'])}種 / "
          f"型式{len(catalog['型式'])}種 / 作業{len(catalog['作業'])}種")


if __name__ == "__main__":
    raise SystemExit(main())
