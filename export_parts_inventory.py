"""部品在庫表.xlsx から在庫データを抽出し、parts_inventory.json を lw-raw Blob に出力する。

目的:
    本番の検索アプリ(Azure App Service)は社内NAS(Z:)を直接見られないため、
    部品在庫を JSON 化して Blob 経由で渡す。検索アプリはこれを読んで
    部品在庫検索を提供する。（工番マスタ export_workno_master.py と同じ方式）

実行:
    Z: が見える環境(デスクトップ)で、在庫表が更新されたら実行する。
    python export_parts_inventory.py

入力 Excel(既定; 環境変数 PARTS_XLSX_PATH で上書き可):
    Z:\takachiho\2to9_業務別フォルダ\95_資材・調達\952_資材管理\
        9521_貯蔵品在庫表(寄居・綾瀬)\部品在庫表(寄居・綾瀬).xlsx

出力 Blob:
    lw-raw/parts_inventory.json
    形式: {"generated_at": ISO8601, "count": N,
           "categories": [...],
           "items": [{"cat","tana","model","spec","maker","supplier","qty","price"}, ...]}

環境変数(.env):
    AZURE_BLOB_CONNECTION_STRING  Blob 接続文字列
    LW_BLOB_CONTAINER             コンテナ名(省略時: lw-raw)
    PARTS_XLSX_PATH               在庫表xlsxのパス(省略時: 上記の既定)
"""
from __future__ import annotations

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).with_name(".env"), encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    print("[ERROR] pip install openpyxl が必要です", file=sys.stderr)
    sys.exit(1)

try:
    from azure.storage.blob import BlobServiceClient
except ImportError:
    print("[ERROR] pip install azure-storage-blob python-dotenv", file=sys.stderr)
    sys.exit(1)

BLOB_CONN_STR = os.environ.get("AZURE_BLOB_CONNECTION_STRING", "")
BLOB_CONTAINER = os.environ.get("LW_BLOB_CONTAINER", "lw-raw")
PARTS_BLOB = "parts_inventory.json"

DEFAULT_XLSX = (
    r"Z:\takachiho\2to9_業務別フォルダ\95_資材・調達\952_資材管理"
    r"\9521_貯蔵品在庫表(寄居・綾瀬)\部品在庫表(寄居・綾瀬).xlsx"
)
XLSX_PATH = os.environ.get("PARTS_XLSX_PATH", DEFAULT_XLSX)

# 在庫データではないシート(集計・除却済み・別形式)は除外する
SKIP_SHEETS = {"在庫総額", "除却", "電気A'ssy", "パッキンセットA'ssy"}


def _num(v) -> str:
    """セル値を表示用文字列に。整数は小数点を落とす。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def parse_inventory(xlsx_path: str) -> list[dict]:
    """在庫表xlsxを走査し、部品アイテムのリストを返す。

    各シートは 7行目付近に「棚番/型式/用途・仕様/メーカー/仕入先/数量/仕入単価/在庫額/見積単価」
    のヘッダーがあり、その次行からデータ。列は位置で読む(シート間で並びが統一)。
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    items: list[dict] = []
    for sn in wb.sheetnames:
        if sn in SKIP_SHEETS:
            continue
        ws = wb[sn]
        # ヘッダー行(「棚番」を含む行)を探す
        header_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
            cells = [("" if c is None else str(c)).replace(" ", "").replace("　", "") for c in row]
            if any("棚番" in c for c in cells):
                header_idx = i  # 0-based
                break
        if header_idx is None:
            continue  # 在庫シートではない
        # データはヘッダーの次行から
        for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
            def g(idx: int) -> str:
                return _num(row[idx]) if idx < len(row) else ""

            tana, model, spec = g(0), g(1), g(2)
            maker, supplier = g(3), g(4)
            qty = g(5)
            # 原価(仕入単価=g(6))は表示しない。見積単価(J列=g(9))を採用し、小数は整数へ丸める。
            _pv = row[9] if 9 < len(row) else None
            price = str(int(round(_pv))) if isinstance(_pv, (int, float)) else _num(_pv)
            # 型式も用途も空 → 空行/区切り行としてスキップ
            if not model and not spec:
                continue
            # 小計・合計行を除外
            if any(x in (tana + model) for x in ("小計", "合計", "総計")):
                continue
            items.append({
                "cat": sn,          # カテゴリ(シート名)
                "tana": tana,       # 棚番
                "model": model,     # 型式 / 図番・品名
                "spec": spec,       # 用途・仕様 / 材質
                "maker": maker,     # メーカー / 客先
                "supplier": supplier,  # 仕入先
                "qty": qty,         # 数量
                "quote": price,     # 見積単価（原価=仕入単価は載せない。キーは quote）
            })
    return items


def main() -> None:
    if not BLOB_CONN_STR:
        print("[ERROR] AZURE_BLOB_CONNECTION_STRING が未設定です", file=sys.stderr)
        sys.exit(1)

    xlsx = Path(XLSX_PATH)
    if not xlsx.exists():
        print(f"[ERROR] 在庫表xlsxが見つかりません: {xlsx}", file=sys.stderr)
        print("        Z: 未接続の可能性。中止します(既存のBlobを保護)。", file=sys.stderr)
        sys.exit(2)

    items = parse_inventory(str(xlsx))
    if not items:
        print("[ERROR] 在庫アイテムが0件。シート構成を確認してください。中止します。", file=sys.stderr)
        sys.exit(3)

    categories = sorted({it["cat"] for it in items})
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": len(items),
        "categories": categories,
        "items": items,
    }

    svc = BlobServiceClient.from_connection_string(BLOB_CONN_STR)
    container = svc.get_container_client(BLOB_CONTAINER)
    container.upload_blob(
        PARTS_BLOB,
        json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        overwrite=True,
        content_type="application/json",
    )
    print(
        f"[DONE] 部品在庫をアップロード: {BLOB_CONTAINER}/{PARTS_BLOB} "
        f"({len(items)} 件 / {len(categories)} カテゴリ)"
    )


if __name__ == "__main__":
    main()
